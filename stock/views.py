from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import F, Sum
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse
from datetime import date
import os
import uuid
from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from .models import (
    Producto, TipoProducto, Cliente, Ventas, DetalleVenta,
    ImagenProducto, Chofer, Envio
)



def login_usuario(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            messages.success(request, f'Bienvenido {user.username}')
            return redirect('home')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')

    return render(request, 'auth/login.html')


@login_required
def logout_usuario(request):
    logout(request)
    messages.success(request, 'Sesión cerrada correctamente')
    return redirect('login')


# ==================================
# VISTAS BÁSICAS
# ==================================
@login_required
def home(request):
    productos_bajo_stock = Producto.objects.filter(
        cantidad__lt=F('umbral_alerta')
    ).count()
    
    con_alerta = productos_bajo_stock > 0
    
    # Ventas pendientes de enviar
    ventas_pendientes = Ventas.objects.filter(
        estado__in=['pendiente', 'confirmada']
    ).count()
    
    contexto = {
        'productos_bajo_stock': productos_bajo_stock,
        'con_alerta': con_alerta,
        'ventas_pendientes': ventas_pendientes,
    }
    
    return render(request, 'home.html', contexto)


# ==================================
# PRODUCTOS
# ==================================
@login_required
def lista_productos(request):
    tipos = TipoProducto.objects.all()
    query_nombre = request.GET.get('nombre')
    query_tipo = request.GET.get('tipo')

    productos = Producto.objects.select_related('tipo').all()

    if query_nombre:
        productos = productos.filter(nombre__icontains=query_nombre)
    if query_tipo:
        productos = productos.filter(tipo_id=query_tipo)

    return render(request, 'lista_productos.html', {
        'productos': productos,
        'tipos': tipos,
        'query_nombre': query_nombre,
        'query_tipo': query_tipo,
    })

@login_required
def crear_tipo(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        if nombre:
            TipoProducto.objects.create(nombre=nombre)
            return redirect('lista_productos')
    return render(request, 'crear_tipo.html')

@login_required
def crear_producto(request):
    tipos = TipoProducto.objects.all()

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        tipo_id = request.POST.get('tipo')
        cantidad = request.POST.get('cantidad')
        valor_compra = request.POST.get('valor_compra')  # 🆕 NUEVO
        valor = request.POST.get('valor')
        umbral_alerta = request.POST.get('umbral_alerta', 5)
        
        try:
            umbral_alerta = int(umbral_alerta)
        except (ValueError, TypeError):
            umbral_alerta = 5

        if nombre and tipo_id:
            tipo = TipoProducto.objects.get(id=tipo_id)
            Producto.objects.create(
                nombre=nombre,
                tipo=tipo,
                cantidad=cantidad,
                valor_compra=valor_compra,  # 🆕 NUEVO
                valor=valor,
                umbral_alerta=umbral_alerta
            )
            return redirect('lista_productos')

    return render(request, 'crear_producto.html', {'tipos': tipos})
@login_required
def actualizar_stock(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)

    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad'))
        accion = request.POST.get('accion')

        if accion == 'sumar':
            producto.cantidad += cantidad
        elif accion == 'restar':
            producto.cantidad -= cantidad
            if producto.cantidad < 0:
                producto.cantidad = 0

        producto.save()
        return redirect('lista_productos')

    return render(request, 'actualizar_stock.html', {'producto': producto})

@login_required
def editar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    tipos = TipoProducto.objects.all()

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        tipo_id = request.POST.get('tipo')
        valor = request.POST.get('valor')
        umbral_alerta = request.POST.get('umbral_alerta')
        
        producto.nombre = nombre
        producto.tipo_id = tipo_id
        producto.valor = valor
        
        try:
            if umbral_alerta is not None:
                producto.umbral_alerta = int(umbral_alerta)
        except ValueError:
            messages.error(request, 'El umbral debe ser un número entero.')
            return redirect('editar_producto', producto_id=producto.id)
            
        producto.save()
        messages.success(request, f'Producto "{producto.nombre}" actualizado.')
        return redirect('lista_productos')

    return render(request, 'editar_producto.html', {
        'producto': producto,
        'tipos': tipos
    })

@login_required
def panel_alertas(request):
    productos_con_alerta = Producto.objects.filter(
        cantidad__lt=F('umbral_alerta')
    ).select_related('tipo').order_by('nombre')
    
    contexto = {
        'productos_con_alerta': productos_con_alerta,
        'conteo_alertas': productos_con_alerta.count(),
    }
    
    return render(request, 'panel_alertas.html', contexto)


# ==================================
# CLIENTES
# ==================================
@login_required
def lista_clientes(request):

    # 🔥 Si es superusuario → ve todo
    if request.user.is_superuser:
        clientes = Cliente.objects.all()

    else:
        perfil = request.user.perfilusuario

        # 🔹 Usuario ventas → solo sus clientes
        if perfil.tipo == 'ventas':
            clientes = Cliente.objects.filter(usuario=request.user)

        # 🔹 Administrativo → todos
        elif perfil.tipo == 'administrativo':
            clientes = Cliente.objects.all()

        # 🔹 Otros (camiones)
        else:
            clientes = Cliente.objects.none()

    return render(request, 'lista_clientes.html', {
        'clientes': clientes
    })

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Cliente


@login_required
def crear_cliente(request):
    if request.method == 'POST':
        nombre_completo = request.POST.get('nombre_completo')
        nombre_local = request.POST.get('nombre_local')
        cuil = request.POST.get('cuil')
        email = request.POST.get('email')
        telefono = request.POST.get('telefono')
        direccion = request.POST.get('direccion')

        if nombre_completo:
            Cliente.objects.create(
                usuario=request.user,  # 🔥 ACA SE GUARDA EL USUARIO LOGUEADO
                nombre_completo=nombre_completo,
                nombre_local=nombre_local if nombre_local else None,
                cuil=cuil if cuil else None,
                email=email if email else None,
                telefono=telefono if telefono else None,
                direccion=direccion if direccion else None
            )

            messages.success(request, 'Cliente creado correctamente')
            return redirect('lista_clientes')

    return render(request, 'crear_cliente.html')
@login_required
def editar_cliente(request, cliente_id):

    cliente = get_object_or_404(Cliente, id=cliente_id)

    perfil = getattr(request.user, 'perfilusuario', None)

    # 🔒 Si es ventas, solo puede editar sus propios clientes
    if not request.user.is_superuser and perfil.tipo == 'ventas':
        if cliente.usuario != request.user:
            messages.error(request, "No tenés permiso para editar este cliente.")
            return redirect('lista_clientes')

    # 🔥 Lista de vendedores (solo para admin y superuser)
    vendedores = None
    if request.user.is_superuser or perfil.tipo == 'administrativo':
        vendedores = User.objects.filter(perfilusuario__tipo='ventas')

    if request.method == 'POST':

        cliente.nombre_completo = request.POST.get('nombre_completo')
        cliente.nombre_local = request.POST.get('nombre_local')
        cliente.email = request.POST.get('email') or None
        cliente.telefono = request.POST.get('telefono') or None
        cliente.direccion = request.POST.get('direccion') or None

        # 🔥 Solo admin o superusuario pueden cambiar vendedor
        if request.user.is_superuser or perfil.tipo == 'administrativo':
            nuevo_usuario_id = request.POST.get('usuario')
            if nuevo_usuario_id:
                cliente.usuario = User.objects.get(id=nuevo_usuario_id)

        cliente.save()

        messages.success(request, "Cliente actualizado correctamente.")
        return redirect('lista_clientes')

    return render(request, 'editar_cliente.html', {
        'cliente': cliente,
        'vendedores': vendedores
    })

# ==================================
# 🆕 VENTAS - VENDEDORES
# ==================================
@login_required
def crear_venta(request):
    """
    SIMPLIFICADO: Solo crear venta con productos.
    Sin chofer, sin programación. Eso viene después.
    Guarda el usuario que crea la venta.
    """
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente')
        notas = request.POST.get('notas')

        productos_ids = request.POST.getlist('productos')
        cantidades = request.POST.getlist('cantidades')

        if cliente_id and productos_ids:
            try:
                cliente = Cliente.objects.get(id=cliente_id)

                # 🔹 Crear la venta SIMPLE (pendiente) con usuario creador
                venta = Ventas.objects.create(
                    cliente=cliente,
                    estado='pendiente',
                    notas=notas or '',
                    valor_total=0,
                    usuario_creador=request.user  # 👈 CLAVE
                )

                total_venta = 0

                # 🔹 Agregar productos
                for prod_id, cant in zip(productos_ids, cantidades):
                    if prod_id and cant:
                        producto = Producto.objects.get(id=prod_id)
                        cantidad = int(cant)
                        precio = producto.valor
                        subtotal = precio * cantidad

                        DetalleVenta.objects.create(
                            venta=venta,
                            producto=producto,
                            cantidad=cantidad,
                            precio_unitario=precio,
                            subtotal=subtotal
                        )

                        total_venta += subtotal

                # 🔹 Actualizar total
                venta.valor_total = total_venta
                venta.save(update_fields=['valor_total'])

                messages.success(
                    request,
                    f'✅ Venta #{venta.id} creada por ${total_venta:.2f}. Estado: Pendiente.'
                )
                return redirect('lista_ventas')

            except Exception as e:
                messages.error(request, f'Error al crear venta: {str(e)}')
        else:
            messages.error(request, 'Seleccione un cliente y al menos un producto')

    clientes = Cliente.objects.all().order_by('nombre_completo')
    productos = Producto.objects.all().order_by('nombre')

    return render(request, 'ventas/crear_venta.html', {
        'clientes': clientes,
        'productos': productos,
    })


@login_required
def lista_ventas(request):
    estado = request.GET.get('estado')
    fecha = request.GET.get('fecha')

    ventas = Ventas.objects.none()
    hay_filtros = any([estado, fecha])

    if hay_filtros:
        ventas = Ventas.objects.filter(
            usuario_creador=request.user  # ✅ CAMPO CORRECTO
        )

        if fecha:
            ventas = ventas.filter(fecha_creacion__date=fecha)

        if estado:
            ventas = ventas.filter(estado=estado)

        ventas = ventas.order_by('-fecha_creacion')

    return render(request, 'ventas/lista_ventas.html', {
        'ventas': ventas,
        'estado': estado,
        'fecha': fecha,
        'estados': Ventas.ESTADO_CHOICES,
        'hay_filtros': hay_filtros,
    })

@login_required
def detalle_venta(request, venta_id):
    """Ver detalles de una venta con opción de asignar chofer"""
    venta = get_object_or_404(Ventas, id=venta_id)
    detalles = venta.detalles.select_related('producto').all()
    choferes = Chofer.objects.filter(activo=True).order_by('nombre_completo')
    
    return render(request, 'ventas/detalle_venta.html', {
        'venta': venta,
        'detalles': detalles,
        'choferes': choferes  # 👈 ESTO ES LO IMPORTANTE
    })


@login_required
def actualizar_estado_venta(request, venta_id):
    """
    Cambiar estado: pendiente ↔ confirmada ↔ cancelada
    CON DEVOLUCIÓN DE STOCK SI SE CANCELA UNA VENTA CONFIRMADA
    """
    venta = get_object_or_404(Ventas, id=venta_id)
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        
        # Validar transiciones permitidas
        transiciones_permitidas = {
            'pendiente': ['confirmada', 'cancelada'],
            'confirmada': ['pendiente', 'cancelada', 'enviada'],
            'enviada': ['entregada'],  
            'entregada': [],
            'cancelada': []
        }
        
        if nuevo_estado not in transiciones_permitidas.get(venta.estado, []):
            messages.error(
                request, 
                f'❌ No se puede cambiar de "{venta.get_estado_display()}" a "{dict(Ventas.ESTADO_CHOICES).get(nuevo_estado)}"'
            )
            return redirect('detalle_venta', venta_id=venta_id)
        
        # 🚨 SI SE CANCELA UNA VENTA CONFIRMADA, DEVOLVER STOCK
        if nuevo_estado == 'cancelada' and venta.estado == 'confirmada':
            try:
                detalles = venta.detalles.select_related('producto').all()
                for detalle in detalles:
                    producto = detalle.producto
                    producto.cantidad = F('cantidad') + detalle.cantidad
                    producto.save()
                    producto.refresh_from_db()
                
                venta.estado = nuevo_estado
                venta.save()
                messages.success(request, '✅ Venta cancelada y stock devuelto')
                
            except Exception as e:
                messages.error(request, f'❌ Error al devolver stock: {str(e)}')
        
        # 🚨 SI SE VUELVE A PENDIENTE DESDE CONFIRMADA, DEVOLVER STOCK
        elif nuevo_estado == 'pendiente' and venta.estado == 'confirmada':
            try:
                detalles = venta.detalles.select_related('producto').all()
                for detalle in detalles:
                    producto = detalle.producto
                    producto.cantidad = F('cantidad') + detalle.cantidad
                    producto.save()
                    producto.refresh_from_db()
                
                venta.estado = nuevo_estado
                venta.chofer = None  # Quitar chofer asignado
                venta.save()
                messages.success(request, '✅ Venta vuelta a pendiente y stock devuelto')
                
            except Exception as e:
                messages.error(request, f'❌ Error al devolver stock: {str(e)}')
        
        else:
            # Cambio normal de estado
            venta.estado = nuevo_estado
            venta.save()
            messages.success(request, f'✅ Estado actualizado: {venta.get_estado_display()}')
        
        return redirect('detalle_venta', venta_id=venta_id)
    
    return render(request, 'ventas/actualizar_estado.html', {
        'venta': venta,
        'estados': Ventas.ESTADO_CHOICES
    })

# ==================================
# CONSULTAR VENTAS (REPORTES)
# ==================================
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.contrib.auth.models import User

@login_required
def consultar_ventas(request):
    """
    Consultar SOLO ventas ENTREGADAS con análisis de ganancia
    """

    productos = Producto.objects.all()
    clientes = Cliente.objects.all().order_by('nombre_completo')

    # ✅ SOLO USUARIOS DE VENTAS
    usuarios = User.objects.filter(
        is_active=True,
        perfilusuario__tipo="ventas"
    ).order_by('username')

    fecha_desde = request.GET.get('desde')
    fecha_hasta = request.GET.get('hasta')
    cliente_id = request.GET.get('cliente')
    producto_id = request.GET.get('producto')
    usuario_id = request.GET.get('usuario')
    exportar = request.GET.get('exportar')

    # 🎯 SOLO ENTREGADAS
    ventas = Ventas.objects.filter(
        estado='entregada'
    ).select_related(
        'cliente',
        'usuario_creador'
    ).prefetch_related(
        'detalles__producto'
    )

    if fecha_desde:
        ventas = ventas.filter(fecha_envio__date__gte=fecha_desde)
    if fecha_hasta:
        ventas = ventas.filter(fecha_envio__date__lte=fecha_hasta)
    if cliente_id:
        ventas = ventas.filter(cliente_id=cliente_id)
    if producto_id:
        ventas = ventas.filter(detalles__producto_id=producto_id).distinct()
    if usuario_id:
        ventas = ventas.filter(usuario_creador_id=usuario_id)

    ventas = ventas.order_by('-fecha_envio')

    # 🧮 Totales
    total_ventas = 0
    total_costo = 0

    for venta in ventas:
        for detalle in venta.detalles.all():
            total_ventas += detalle.subtotal
            total_costo += (detalle.producto.valor_compra or 0) * detalle.cantidad

    ganancia_total = total_ventas - total_costo

    # 📊 Exportar
    if exportar == 'excel':
        return exportar_ventas_excel(ventas, fecha_desde, fecha_hasta)

    return render(request, 'consultar_ventas.html', {
        'ventas': ventas,
        'productos': productos,
        'clientes': clientes,
        'usuarios': usuarios,  # 👈 ahora SOLO ventas
        'total_ventas': total_ventas,
        'total_costo': total_costo,
        'ganancia_total': ganancia_total,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'cliente_id': cliente_id,
        'producto_id': producto_id,
        'usuario_id': usuario_id,
    })

def exportar_ventas_excel(ventas, fecha_desde, fecha_hasta):
    """
    Exporta las ventas a Excel con análisis detallado
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"
    
    # 🎨 Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=14)
    
    # 📋 Título
    ws.merge_cells('A1:J1')
    ws['A1'] = f'REPORTE DE VENTAS ENTREGADAS'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    if fecha_desde or fecha_hasta:
        ws.merge_cells('A2:J2')
        periodo = f"Periodo: {fecha_desde or 'Inicio'} hasta {fecha_hasta or 'Hoy'}"
        ws['A2'] = periodo
        ws['A2'].alignment = Alignment(horizontal='center')
        fila_inicio = 4
    else:
        fila_inicio = 3
    
    # 📊 Encabezados
    headers = [
        'Fecha Entrega', 'Venta #', 'Cliente', 'Vendedor', 'Producto', 
        'Cantidad', 'Precio Compra Unit.', 'Precio Venta Unit.', 
        'Costo Total', 'Venta Total', 'Ganancia'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=fila_inicio, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # 📝 Datos
    fila = fila_inicio + 1
    total_costo = 0
    total_venta = 0
    total_ganancia = 0
    
    for venta in ventas:
        for detalle in venta.detalles.all():
            costo_unit = detalle.producto.valor_compra or 0
            precio_unit = detalle.precio_unitario
            cantidad = detalle.cantidad
            
            costo_total = costo_unit * cantidad
            venta_total = detalle.subtotal
            ganancia = venta_total - costo_total
            
            ws.cell(row=fila, column=1).value = venta.fecha_envio.strftime('%d/%m/%Y %H:%M') if venta.fecha_envio else 'N/A'
            ws.cell(row=fila, column=2).value = venta.id
            ws.cell(row=fila, column=3).value = venta.cliente.nombre_completo
            ws.cell(row=fila, column=4).value = venta.usuario_creador.username if venta.usuario_creador else 'N/A'
            ws.cell(row=fila, column=5).value = detalle.producto.nombre
            ws.cell(row=fila, column=6).value = cantidad
            ws.cell(row=fila, column=7).value = float(costo_unit)
            ws.cell(row=fila, column=8).value = float(precio_unit)
            ws.cell(row=fila, column=9).value = float(costo_total)
            ws.cell(row=fila, column=10).value = float(venta_total)
            ws.cell(row=fila, column=11).value = float(ganancia)
            
            # Formato moneda
            for col in [7, 8, 9, 10, 11]:
                ws.cell(row=fila, column=col).number_format = '$#,##0.00'
            
            total_costo += costo_total
            total_venta += venta_total
            total_ganancia += ganancia
            
            fila += 1
    
    # 📊 TOTALES
    fila += 1
    ws.merge_cells(f'A{fila}:H{fila}')
    ws.cell(row=fila, column=1).value = 'TOTALES'
    ws.cell(row=fila, column=1).font = Font(bold=True, size=12)
    ws.cell(row=fila, column=9).value = float(total_costo)
    ws.cell(row=fila, column=10).value = float(total_venta)
    ws.cell(row=fila, column=11).value = float(total_ganancia)
    
    for col in [9, 10, 11]:
        cell = ws.cell(row=fila, column=col)
        cell.font = Font(bold=True, size=12)
        cell.number_format = '$#,##0.00'
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # 📏 Ajustar anchos
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    
    # 💾 Generar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'ventas_entregadas_{fecha_desde or "inicio"}_{fecha_hasta or "hoy"}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
# ==================================
# CHOFERES
# ==================================
@login_required
def lista_choferes(request):
    choferes = Chofer.objects.all()
    return render(request, 'envios/lista_choferes.html', {'choferes': choferes})

@login_required
def crear_chofer(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre_completo')
        telefono = request.POST.get('telefono')
        vehiculo = request.POST.get('vehiculo')
        pin = request.POST.get('pin')
        notas = request.POST.get('notas')

        if nombre and telefono and vehiculo and pin:
            Chofer.objects.create(
                nombre_completo=nombre,
                telefono=telefono,
                vehiculo=vehiculo,
                pin=pin,
                notas=notas
            )
            messages.success(request, 'Chofer creado correctamente')
            return redirect('lista_choferes')
        else:
            messages.error(request, 'Complete todos los campos obligatorios')

    return render(request, 'envios/crear_chofer.html')

@login_required
def editar_chofer(request, chofer_id):
    chofer = get_object_or_404(Chofer, id=chofer_id)
    
    if request.method == 'POST':
        chofer.nombre_completo = request.POST.get('nombre_completo')
        chofer.telefono = request.POST.get('telefono')
        chofer.vehiculo = request.POST.get('vehiculo')
        chofer.notas = request.POST.get('notas')
        chofer.activo = request.POST.get('activo') == 'on'
        chofer.save()
        
        messages.success(request, f'Chofer "{chofer.nombre_completo}" actualizado')
        return redirect('lista_choferes')
    
    return render(request, 'envios/editar_chofer.html', {'chofer': chofer})


# ==================================
# 🆕 VENTAS - VENDEDORES
# ==================================
 
@login_required
def lista_ventas(request):
    estado = request.GET.get('estado')
    fecha = request.GET.get('fecha')

    ventas = Ventas.objects.none()
    hay_filtros = any([estado, fecha])

    if hay_filtros:
        ventas = Ventas.objects.filter(
            usuario_creador=request.user  # ✅ CAMPO CORRECTO
        )

        if fecha:
            ventas = ventas.filter(fecha_creacion__date=fecha)

        if estado:
            ventas = ventas.filter(estado=estado)

        ventas = ventas.order_by('-fecha_creacion')

    return render(request, 'ventas/lista_ventas.html', {
        'ventas': ventas,
        'estado': estado,
        'fecha': fecha,
        'estados': Ventas.ESTADO_CHOICES,
        'hay_filtros': hay_filtros,
    })

@login_required
def detalle_venta(request, venta_id):
    """Ver detalles de una venta con opción de asignar chofer"""
    venta = get_object_or_404(Ventas, id=venta_id)
    detalles = venta.detalles.select_related('producto').all()
    choferes = Chofer.objects.filter(activo=True).order_by('nombre_completo')
    
    return render(request, 'ventas/detalle_venta.html', {
        'venta': venta,
        'detalles': detalles,
        'choferes': choferes  # 👈 ESTO ES LO IMPORTANTE
    })


@login_required
def actualizar_estado_venta(request, venta_id):
    """
    Cambiar estado: pendiente ↔ confirmada ↔ cancelada
    CON DEVOLUCIÓN DE STOCK SI SE CANCELA UNA VENTA CONFIRMADA
    """
    venta = get_object_or_404(Ventas, id=venta_id)
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        
        # Validar transiciones permitidas
        transiciones_permitidas = {
            'pendiente': ['confirmada', 'cancelada'],
            'confirmada': ['pendiente', 'cancelada', 'enviada'],
            'enviada': ['entregada'],  
            'entregada': [],
            'cancelada': []
        }
        
        if nuevo_estado not in transiciones_permitidas.get(venta.estado, []):
            messages.error(
                request, 
                f'❌ No se puede cambiar de "{venta.get_estado_display()}" a "{dict(Ventas.ESTADO_CHOICES).get(nuevo_estado)}"'
            )
            return redirect('detalle_venta', venta_id=venta_id)
        
        # 🚨 SI SE CANCELA UNA VENTA CONFIRMADA, DEVOLVER STOCK
        if nuevo_estado == 'cancelada' and venta.estado == 'confirmada':
            try:
                detalles = venta.detalles.select_related('producto').all()
                for detalle in detalles:
                    producto = detalle.producto
                    producto.cantidad = F('cantidad') + detalle.cantidad
                    producto.save()
                    producto.refresh_from_db()
                
                venta.estado = nuevo_estado
                venta.save()
                messages.success(request, '✅ Venta cancelada y stock devuelto')
                
            except Exception as e:
                messages.error(request, f'❌ Error al devolver stock: {str(e)}')
        
        # 🚨 SI SE VUELVE A PENDIENTE DESDE CONFIRMADA, DEVOLVER STOCK
        elif nuevo_estado == 'pendiente' and venta.estado == 'confirmada':
            try:
                detalles = venta.detalles.select_related('producto').all()
                for detalle in detalles:
                    producto = detalle.producto
                    producto.cantidad = F('cantidad') + detalle.cantidad
                    producto.save()
                    producto.refresh_from_db()
                
                venta.estado = nuevo_estado
                venta.chofer = None  # Quitar chofer asignado
                venta.save()
                messages.success(request, '✅ Venta vuelta a pendiente y stock devuelto')
                
            except Exception as e:
                messages.error(request, f'❌ Error al devolver stock: {str(e)}')
        
        else:
            # Cambio normal de estado
            venta.estado = nuevo_estado
            venta.save()
            messages.success(request, f'✅ Estado actualizado: {venta.get_estado_display()}')
        
        return redirect('detalle_venta', venta_id=venta_id)
    
    return render(request, 'ventas/actualizar_estado.html', {
        'venta': venta,
        'estados': Ventas.ESTADO_CHOICES
    })

# ==================================
# CONSULTAR VENTAS (REPORTES)
# ==================================
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.contrib.auth.models import User

@login_required
def consultar_ventas(request):
    """
    Consultar SOLO ventas ENTREGADAS con análisis de ganancia
    """

    productos = Producto.objects.all()
    clientes = Cliente.objects.all().order_by('nombre_completo')

    # ✅ SOLO USUARIOS DE VENTAS
    usuarios = User.objects.filter(
        is_active=True,
        perfilusuario__tipo="ventas"
    ).order_by('username')

    fecha_desde = request.GET.get('desde')
    fecha_hasta = request.GET.get('hasta')
    cliente_id = request.GET.get('cliente')
    producto_id = request.GET.get('producto')
    usuario_id = request.GET.get('usuario')
    exportar = request.GET.get('exportar')

    # 🎯 SOLO ENTREGADAS
    ventas = Ventas.objects.filter(
        estado='entregada'
    ).select_related(
        'cliente',
        'usuario_creador'
    ).prefetch_related(
        'detalles__producto'
    )

    if fecha_desde:
        ventas = ventas.filter(fecha_envio__date__gte=fecha_desde)
    if fecha_hasta:
        ventas = ventas.filter(fecha_envio__date__lte=fecha_hasta)
    if cliente_id:
        ventas = ventas.filter(cliente_id=cliente_id)
    if producto_id:
        ventas = ventas.filter(detalles__producto_id=producto_id).distinct()
    if usuario_id:
        ventas = ventas.filter(usuario_creador_id=usuario_id)

    ventas = ventas.order_by('-fecha_envio')

    # 🧮 Totales
    total_ventas = 0
    total_costo = 0

    for venta in ventas:
        for detalle in venta.detalles.all():
            total_ventas += detalle.subtotal
            total_costo += (detalle.producto.valor_compra or 0) * detalle.cantidad

    ganancia_total = total_ventas - total_costo

    # 📊 Exportar
    if exportar == 'excel':
        return exportar_ventas_excel(ventas, fecha_desde, fecha_hasta)

    return render(request, 'consultar_ventas.html', {
        'ventas': ventas,
        'productos': productos,
        'clientes': clientes,
        'usuarios': usuarios,  # 👈 ahora SOLO ventas
        'total_ventas': total_ventas,
        'total_costo': total_costo,
        'ganancia_total': ganancia_total,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'cliente_id': cliente_id,
        'producto_id': producto_id,
        'usuario_id': usuario_id,
    })

def exportar_ventas_excel(ventas, fecha_desde, fecha_hasta):
    """
    Exporta las ventas a Excel con análisis detallado
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"
    
    # 🎨 Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=14)
    
    # 📋 Título
    ws.merge_cells('A1:J1')
    ws['A1'] = f'REPORTE DE VENTAS ENTREGADAS'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    if fecha_desde or fecha_hasta:
        ws.merge_cells('A2:J2')
        periodo = f"Periodo: {fecha_desde or 'Inicio'} hasta {fecha_hasta or 'Hoy'}"
        ws['A2'] = periodo
        ws['A2'].alignment = Alignment(horizontal='center')
        fila_inicio = 4
    else:
        fila_inicio = 3
    
    # 📊 Encabezados
    headers = [
        'Fecha Entrega', 'Venta #', 'Cliente', 'Vendedor', 'Producto', 
        'Cantidad', 'Precio Compra Unit.', 'Precio Venta Unit.', 
        'Costo Total', 'Venta Total', 'Ganancia'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=fila_inicio, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # 📝 Datos
    fila = fila_inicio + 1
    total_costo = 0
    total_venta = 0
    total_ganancia = 0
    
    for venta in ventas:
        for detalle in venta.detalles.all():
            costo_unit = detalle.producto.valor_compra or 0
            precio_unit = detalle.precio_unitario
            cantidad = detalle.cantidad
            
            costo_total = costo_unit * cantidad
            venta_total = detalle.subtotal
            ganancia = venta_total - costo_total
            
            ws.cell(row=fila, column=1).value = venta.fecha_envio.strftime('%d/%m/%Y %H:%M') if venta.fecha_envio else 'N/A'
            ws.cell(row=fila, column=2).value = venta.id
            ws.cell(row=fila, column=3).value = venta.cliente.nombre_completo
            ws.cell(row=fila, column=4).value = venta.usuario_creador.username if venta.usuario_creador else 'N/A'
            ws.cell(row=fila, column=5).value = detalle.producto.nombre
            ws.cell(row=fila, column=6).value = cantidad
            ws.cell(row=fila, column=7).value = float(costo_unit)
            ws.cell(row=fila, column=8).value = float(precio_unit)
            ws.cell(row=fila, column=9).value = float(costo_total)
            ws.cell(row=fila, column=10).value = float(venta_total)
            ws.cell(row=fila, column=11).value = float(ganancia)
            
            # Formato moneda
            for col in [7, 8, 9, 10, 11]:
                ws.cell(row=fila, column=col).number_format = '$#,##0.00'
            
            total_costo += costo_total
            total_venta += venta_total
            total_ganancia += ganancia
            
            fila += 1
    
    # 📊 TOTALES
    fila += 1
    ws.merge_cells(f'A{fila}:H{fila}')
    ws.cell(row=fila, column=1).value = 'TOTALES'
    ws.cell(row=fila, column=1).font = Font(bold=True, size=12)
    ws.cell(row=fila, column=9).value = float(total_costo)
    ws.cell(row=fila, column=10).value = float(total_venta)
    ws.cell(row=fila, column=11).value = float(total_ganancia)
    
    for col in [9, 10, 11]:
        cell = ws.cell(row=fila, column=col)
        cell.font = Font(bold=True, size=12)
        cell.number_format = '$#,##0.00'
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # 📏 Ajustar anchos
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    
    # 💾 Generar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'ventas_entregadas_{fecha_desde or "inicio"}_{fecha_hasta or "hoy"}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

# ==================================
# ENVÍOS (OPCIONAL - TRACKING)
# ==================================
def lista_envios(request):
    """
    Lista envíos con filtros opcionales y al mismo tiempo permite filtrar por ventas.
    Trabaja con el modelo Envio, que está vinculado a Ventas.
    """
    fecha = request.GET.get('fecha')
    chofer_id = request.GET.get('chofer')
    estado_envio = request.GET.get('estado_envio')  # Filtrar por estado del envío
    estado_venta = request.GET.get('estado_venta')  # Filtrar por estado de la venta

    # ⚠️ No cargar nada por defecto
    envios = Envio.objects.select_related('venta__cliente', 'chofer')

    # Filtrar si hay parámetros
    if fecha:
        envios = envios.filter(fecha_envio=fecha)

    if chofer_id:
        envios = envios.filter(chofer_id=chofer_id)

    if estado_envio:
        envios = envios.filter(estado=estado_envio)

    if estado_venta:
        envios = envios.filter(venta__estado=estado_venta)

    envios = envios.order_by('hora_estimada')

    choferes = Chofer.objects.filter(activo=True)

    context = {
        'envios': envios,
        'choferes': choferes,
        'fecha': fecha,
        'chofer_id': chofer_id,
        'estado_envio': estado_envio,
        'estado_venta': estado_venta,
        'estados_envio': Envio.ESTADO_CHOICES,
        'estados_venta': Ventas.ESTADO_CHOICES,
        'total_envios': envios.count(),
        'envios_pendientes': envios.filter(estado='pendiente').count(),
        'envios_en_camino': envios.filter(estado='en_camino').count(),
        'envios_entregados': envios.filter(estado='entregado').count(),
        'hay_filtros': bool(fecha or chofer_id or estado_envio or estado_venta),
    }

    return render(request, 'envios/lista_envios.html', context)

@login_required
def crear_envio(request, venta_id):
    """
    Crear envío = solo marcar como ENVIADA (el stock YA fue descontado)
    """
    venta = get_object_or_404(Ventas, id=venta_id)
    
    # Verificar que no tenga envío ya
    if hasattr(venta, 'envio'):
        messages.warning(request, 'Esta venta ya tiene un envío asignado')
        return redirect('lista_envios')
    
    # Solo ventas confirmadas
    if venta.estado != 'confirmada':
        messages.error(
            request, 
            f'Solo ventas confirmadas pueden crear envío. Estado actual: {venta.get_estado_display()}'
        )
        return redirect('detalle_venta', venta_id=venta_id)
    
    if request.method == 'POST':
        chofer_id = request.POST.get('chofer')
        fecha_envio = request.POST.get('fecha_envio')
        hora_estimada = request.POST.get('hora_estimada')
        direccion = request.POST.get('direccion_entrega')
        notas = request.POST.get('notas')
        
        if chofer_id and fecha_envio and hora_estimada:
            try:
                chofer = Chofer.objects.get(id=chofer_id)
                
                # 🎯 EL STOCK YA FUE DESCONTADO EN asignar_chofer_venta
                # 🎯 Solo crear el envío
                Envio.objects.create(
                    venta=venta,
                    chofer=chofer,
                    fecha_envio=fecha_envio,
                    hora_estimada=hora_estimada,
                    direccion_entrega=direccion or venta.cliente.direccion or '',
                    notas=notas or '',
                    estado='pendiente'
                )
                
                # 🎯 Marcar venta como ENVIADA
                venta.estado = 'enviada'
                venta.fecha_envio = timezone.now()
                venta.save()
                
                messages.success(
                    request, 
                    f'✅ Envío creado. Venta #{venta.id} marcada como ENVIADA.'
                )
                return redirect('lista_envios')
                
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
        else:
            messages.error(request, 'Complete todos los campos obligatorios')
    
    choferes = Chofer.objects.filter(activo=True).order_by('nombre_completo')
    
    return render(request, 'envios/crear_envio.html', {
        'venta': venta,
        'choferes': choferes,
        'fecha_hoy': date.today()
    })

@login_required
def detalle_envio(request, envio_id):
    """Ver detalles de un envío"""
    envio = get_object_or_404(Envio, id=envio_id)
    
    return render(request, 'envios/detalle_envio.html', {
        'envio': envio,
    })

@login_required
def actualizar_estado_envio(request, envio_id):
    """
    Actualizar estado del envío (para choferes)
    """
    envio = get_object_or_404(Envio, id=envio_id)
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        
        if nuevo_estado not in dict(Envio.ESTADO_CHOICES):
            messages.error(request, 'Estado inválido')
            return redirect('detalle_envio', envio_id=envio_id)
        
        # Actualizar estado del envío
        envio.estado = nuevo_estado
        
        # Si se marca como entregado, registrar hora
        if nuevo_estado == 'entregado' and not envio.hora_real_entrega:
            envio.hora_real_entrega = timezone.now()
            
            # Actualizar también el estado de la venta a "entregada"
            if envio.venta.estado == 'enviada':
                envio.venta.estado = 'entregada'
                envio.venta.save()
                messages.success(request, f'✅ Envío entregado. Venta #{envio.venta.id} marcada como entregada.')
        
        envio.save()
        
        if nuevo_estado != 'entregado':
            messages.success(request, f'Estado del envío actualizado a "{envio.get_estado_display()}"')
        
        return redirect('detalle_envio', envio_id=envio_id)
    
    return render(request, 'envios/actualizar_estado.html', {
        'envio': envio,
        'estados': Envio.ESTADO_CHOICES
    })

@login_required
def programa_dia(request):
    """Vista del programa diario de envíos por chofer"""
    fecha = request.GET.get('fecha', date.today())
    
    envios = Envio.objects.filter(fecha_envio=fecha).select_related(
        'venta__cliente', 'chofer'
    ).order_by('chofer', 'hora_estimada')
    
    # Agrupar por chofer
    envios_por_chofer = {}
    for envio in envios:
        chofer_nombre = envio.chofer.nombre_completo if envio.chofer else "Sin Asignar"
        if chofer_nombre not in envios_por_chofer:
            envios_por_chofer[chofer_nombre] = []
        envios_por_chofer[chofer_nombre].append(envio)
    
    return render(request, 'envios/programa_dia.html', {
        'envios_por_chofer': envios_por_chofer,
        'fecha': fecha
    })


# ==================================
# IMÁGENES
# ==================================
@login_required
def subir_imagen(request):
    productos = Producto.objects.all()
    
    if request.method == 'POST':
        producto_id = request.POST.get('producto_id')
        imagen = request.FILES.get('imagen')
        
        if producto_id and imagen:
            try:
                producto = Producto.objects.get(id=producto_id)
                nombre_archivo = f"{uuid.uuid4()}{os.path.splitext(imagen.name)[1]}"
                ruta_relativa = f"productos/{nombre_archivo}"
                ruta_completa = os.path.join(settings.MEDIA_ROOT, ruta_relativa)
                
                os.makedirs(os.path.dirname(ruta_completa), exist_ok=True)
                
                with open(ruta_completa, 'wb+') as destino:
                    for chunk in imagen.chunks():
                        destino.write(chunk)
                
                producto.imagenproducto_set.all().delete()
                ImagenProducto.objects.create(producto=producto, ruta=ruta_relativa)
                
                messages.success(request, f'Imagen subida correctamente para {producto.nombre}')
            except Producto.DoesNotExist:
                messages.error(request, 'Producto no encontrado')
            except Exception as e:
                messages.error(request, f'Error al subir la imagen: {str(e)}')
        else:
            messages.error(request, 'Debe seleccionar un producto y una imagen')
        
        return redirect('subir_imagen')
    
    return render(request, 'imagenes/subir.html', {'productos': productos})

def api_productos(request):
    productos = Producto.objects.all()
    
    lista = []
    for p in productos:
        img = p.imagenproducto_set.first()
        imagen_url = None
        if img and img.ruta:
            imagen_url = request.build_absolute_uri(f'{settings.MEDIA_URL}{img.ruta}')
        
        lista.append({
            'id': p.id,
            'nombre': p.nombre,
            'valor': float(p.valor),
            'imagen': imagen_url
        })
    
    return JsonResponse(lista, safe=False)

@login_required
def asignar_chofer_venta(request, venta_id):
    """
    Admin: asigna un chofer a una venta pendiente y DESCUENTA STOCK
    """
    venta = get_object_or_404(Ventas, id=venta_id, estado__in=['pendiente', 'confirmada'])

    if request.method == 'POST':
        chofer_id = request.POST.get('chofer')
        
        if not chofer_id:
            messages.error(request, "Debes seleccionar un chofer")
            return redirect('detalle_venta', venta_id=venta.id)
        
        try:
            chofer = Chofer.objects.get(id=chofer_id, activo=True)
        except Chofer.DoesNotExist:
            messages.error(request, "❌ El chofer no existe o no está activo")
            return redirect('detalle_venta', venta_id=venta.id)
        
        # 🚨 DESCONTAR STOCK SI ES PENDIENTE
        if venta.estado == 'pendiente':
            try:
                detalles = venta.detalles.select_related('producto').all()
                
                # 1. Verificar stock primero
                for detalle in detalles:
                    if detalle.producto.cantidad < detalle.cantidad:
                        messages.error(
                            request,
                            f"❌ Stock insuficiente de '{detalle.producto.nombre}'. "
                            f"Disponible: {detalle.producto.cantidad}, Necesario: {detalle.cantidad}"
                        )
                        return redirect('detalle_venta', venta_id=venta.id)
                
                # 2. Descontar stock
                for detalle in detalles:
                    producto = detalle.producto
                    producto.cantidad = F('cantidad') - detalle.cantidad
                    producto.save()
                    producto.refresh_from_db()  # Actualizar
                
                messages.success(request, "✅ Stock descontado correctamente")
                
            except Exception as e:
                messages.error(request, f"❌ Error al descontar stock: {str(e)}")
                return redirect('detalle_venta', venta_id=venta.id)
        
        # Actualizar venta
        venta.chofer = chofer
        venta.estado = 'confirmada'  # Siempre pasa a confirmada
        
        # Guardar fechas
        fecha = request.POST.get('fecha_envio_programada')
        hora = request.POST.get('hora_envio_programada')
        
        if fecha:
            venta.fecha_envio_programada = fecha
        if hora:
            venta.hora_envio_programada = hora
        
        venta.save()
        
        messages.success(request, f"✅ Venta #{venta.id} confirmada y asignada a {chofer.nombre_completo}")
        return redirect('detalle_venta', venta_id=venta.id)

    choferes = Chofer.objects.filter(activo=True)
    context = {
        'venta': venta,
        'choferes': choferes,
    }
    return render(request, 'envios/asignar_chofer.html', context)

@login_required
def asignar_envios_pendientes(request):
    """
    Admin: Muestra ventas PENDIENTES y CONFIRMADAS
    Permite asignar un chofer a ventas pendientes o confirmar ventas
    """
    cliente_id = request.GET.get('cliente')
    fecha = request.GET.get('fecha')

    # Traer todas las ventas pendientes y confirmadas
    ventas = Ventas.objects.filter(
        estado__in=['pendiente', 'confirmada']
    ).select_related('cliente', 'chofer').prefetch_related('detalles__producto')

    if cliente_id:
        ventas = ventas.filter(cliente_id=cliente_id)
    if fecha:
        ventas = ventas.filter(fecha_creacion__date=fecha)

    ventas = ventas.order_by('estado', '-fecha_creacion')

    clientes = Cliente.objects.all().order_by('nombre_completo')
    choferes = Chofer.objects.filter(activo=True).order_by('nombre_completo')

    # Separar por estado para la plantilla
    ventas_pendientes = ventas.filter(estado='pendiente')
    ventas_confirmadas = ventas.filter(estado='confirmada')

    context = {
        'ventas_pendientes': ventas_pendientes,
        'ventas_confirmadas': ventas_confirmadas,
        'clientes': clientes,
        'choferes': choferes,
        'cliente_id': cliente_id,
        'fecha': fecha,
    }

    return render(request, 'envios/asignar_envios_pendientes.html', context)

@login_required
def confirmar_y_crear_envio(request, venta_id):
    """
    🆕 NUEVA FUNCIÓN: Confirma la venta automáticamente y va a crear envío
    """
    venta = get_object_or_404(Ventas, id=venta_id)
    
    # Si está pendiente, la confirmamos automáticamente
    if venta.estado == 'pendiente':
        venta.estado = 'confirmada'
        venta.save()
        messages.success(request, f'✅ Venta #{venta.id} confirmada automáticamente')
    
    # Redirigir a crear envío
    return redirect('crear_envio', venta_id=venta_id)


# Agregar estas vistas al archivo views.py existente

# ==================================
# 🚚 PANEL DE CHOFERES
# ==================================
@login_required
def panel_chofer(request):
    """
    Panel del chofer: muestra ventas CONFIRMADAS y ENTREGADAS
    que le fueron asignadas.
    """
    chofer_id = request.session.get('chofer_id')
    if not chofer_id:
        return redirect('acceso_chofer')  # Login de chofer

    chofer = get_object_or_404(Chofer, id=chofer_id, activo=True)

    # Ventas asignadas: confirmadas o entregadas
    ventas_asignadas = Ventas.objects.filter(
        chofer=chofer,
        estado__in=['confirmada', 'enviada']
    ).select_related('cliente').prefetch_related('detalles__producto').order_by('fecha_envio_programada')

    context = {
        'chofer': chofer,
        'ventas_asignadas': ventas_asignadas,
    }

    return render(request, 'choferes/panel_chofer.html', context)



def chofer_cerrar_sesion(request):
    request.session.pop('chofer_id', None)
    return redirect('acceso_chofer')


def acceso_chofer(request):
    choferes = Chofer.objects.filter(activo=True).order_by('nombre_completo')

    if request.method == 'POST':
        chofer_id = request.POST.get('chofer_id')
        pin = request.POST.get('pin')

        chofer = get_object_or_404(Chofer, id=chofer_id, activo=True)

        if pin == chofer.pin:
            request.session['chofer_id'] = chofer.id
            return redirect('panel_chofer')
        else:
            messages.error(request, 'PIN incorrecto')

    return render(request, 'choferes/acceso_chofer.html', {
        'choferes': choferes
    })


@login_required
def chofer_detalle_venta_confirmada(request, venta_id):
    chofer_id = request.session.get('chofer_id')
    if not chofer_id:
        return redirect('acceso_chofer')

    venta = get_object_or_404(Ventas, id=venta_id, chofer_id=chofer_id)

    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        notas = request.POST.get('notas_adicionales')

        # Validar estado
        estados_validos = [e[0] for e in Ventas.ESTADO_CHOICES]
        if nuevo_estado not in estados_validos:
            messages.error(request, 'Estado inválido')
            return redirect('chofer_detalle_venta_confirmada', venta_id=venta_id)

        # Actualizar estado
        venta.estado = nuevo_estado

        # Guardar notas
        if notas:
            timestamp = timezone.now().strftime('%d/%m/%Y %H:%M')
            if venta.notas:
                venta.notas += f"\n[{timestamp}] {notas}"
            else:
                venta.notas = f"[{timestamp}] {notas}"

        venta.save()
        messages.success(request, f'✅ Estado actualizado: {venta.get_estado_display()}')
        return redirect('panel_chofer')

    detalles = venta.detalles.select_related('producto').all()
    return render(request, 'choferes/detalle_venta_confirmada.html', {
        'venta': venta,
        'detalles': detalles,
        'chofer': venta.chofer,
        'estados': Ventas.ESTADO_CHOICES
    })
@login_required
def chofer_detalle_envio(request, envio_id):
    """
    Detalle de un ENVÍO (ya creado)
    """
    chofer_id = request.session.get('chofer_id')
    if not chofer_id:
        return redirect('panel_chofer')
    
    envio = get_object_or_404(Envio, id=envio_id, chofer_id=chofer_id)
    detalles = envio.venta.detalles.select_related('producto').all()
    
    return render(request, 'choferes/detalle_envio.html', {
        'envio': envio,
        'detalles': detalles,
        'chofer': envio.chofer,
        'estados': Envio.ESTADO_CHOICES,
    })

@login_required
def chofer_cambiar_estado_envio(request, envio_id):
    """
    Permite al chofer cambiar el estado de su ENVÍO
    """
    chofer_id = request.session.get('chofer_id')
    if not chofer_id:
        return redirect('panel_chofer')
    
    envio = get_object_or_404(Envio, id=envio_id, chofer_id=chofer_id)
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        notas_adicionales = request.POST.get('notas_adicionales')
        
        # Validar que el estado sea válido
        estados_validos = dict(Envio.ESTADO_CHOICES).keys()
        if nuevo_estado not in estados_validos:
            messages.error(request, 'Estado inválido')
            return redirect('chofer_detalle_envio', envio_id=envio_id)
        
        # Actualizar estado
        envio.estado = nuevo_estado
        
        # Si se marca como entregado, registrar la hora
        if nuevo_estado == 'entregado' and not envio.hora_real_entrega:
            envio.hora_real_entrega = timezone.now()
            
            # Actualizar también la venta a "entregada"
            if envio.venta.estado == 'enviada':
                envio.venta.estado = 'entregada'
                envio.venta.save()
        
        # Agregar notas si las hay
        if notas_adicionales:
            if envio.notas:
                envio.notas += f"\n[{timezone.now().strftime('%d/%m/%Y %H:%M')}] {notas_adicionales}"
            else:
                envio.notas = f"[{timezone.now().strftime('%d/%m/%Y %H:%M')}] {notas_adicionales}"
        
        envio.save()
        
        messages.success(request, f'Estado actualizado a: {envio.get_estado_display()}')
        
        return redirect('chofer_detalle_envio', envio_id=envio_id)
    
    # Si no es POST, mostrar formulario
    return render(request, 'choferes/cambiar_estado.html', {
        'envio': envio,
        'estados': Envio.ESTADO_CHOICES,
    })

@login_required
def chofer_historial(request):
    """
    Historial de envíos del chofer
    """
    chofer_id = request.session.get('chofer_id')
    if not chofer_id:
        return redirect('panel_chofer')
    
    chofer = get_object_or_404(Chofer, id=chofer_id)
    
    # Filtros
    fecha_desde = request.GET.get('desde')
    fecha_hasta = request.GET.get('hasta')
    estado = request.GET.get('estado')
    
    envios = Envio.objects.filter(chofer=chofer).select_related(
        'venta__cliente'
    ).order_by('-fecha_envio', '-hora_estimada')
    
    if fecha_desde:
        envios = envios.filter(fecha_envio__gte=fecha_desde)
    if fecha_hasta:
        envios = envios.filter(fecha_envio__lte=fecha_hasta)
    if estado:
        envios = envios.filter(estado=estado)
    
    # Estadísticas
    total_envios = envios.count()
    total_entregados = envios.filter(estado='entregado').count()
    
    context = {
        'chofer': chofer,
        'envios': envios,
        'total_envios': total_envios,
        'total_entregados': total_entregados,
        'estados': Envio.ESTADO_CHOICES,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'estado': estado,
    }
    
    return render(request, 'choferes/historial.html', context)

@login_required
def chofer_cerrar_sesion(request):
    """
    Cerrar sesión del chofer
    """
    if 'chofer_id' in request.session:
        del request.session['chofer_id']
    messages.success(request, 'Sesión cerrada correctamente')
    return redirect('panel_chofer')